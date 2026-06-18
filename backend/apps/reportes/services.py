from decimal import Decimal
from io import BytesIO
import csv
from datetime import date, datetime
from zoneinfo import ZoneInfo
from django.template.loader import render_to_string
from django.db.models import Sum
from django.http import HttpResponse
from weasyprint import HTML
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from apps.ingresos.models import RegistroNomina, OtroIngreso
from apps.gastos.models import Gasto
from apps.deudas.models import Credito, CuotaCredito, TarjetaCredito, CompraTC
from apps.provisiones.models import Provision, AporteProvision, FondoEmergencia
from apps.indicadores.services import calcular_indicadores_mes

BOGOTA_TZ = ZoneInfo("America/Bogota")
MESES_NOMBRE = [
    "",
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]
HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _marca_temporal() -> str:
    ahora = datetime.now(BOGOTA_TZ)
    return ahora.strftime("%d/%m/%Y %H:%M")


def _nombre_archivo(base: str, formato: str, usuario: str, mes: int, anio: int) -> str:
    ts = datetime.now(BOGOTA_TZ).strftime("%Y%m%d_%H%M%S")
    return f"{base}_{usuario}_{anio}{mes:02d}_{ts}.{formato}"


def _estilo_cabecera(ws, headers, row=1):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER_THIN


def _autoajuste_columnas(ws, min_width=12, max_width=40):
    for col_cells in ws.columns:
        max_len = min_width
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[col_letter].width = max_len + 2


def generar_pdf_mes(usuario, mes: int, anio: int) -> HttpResponse:
    indicadores = calcular_indicadores_mes(usuario, mes, anio)

    gastos = (
        Gasto.objects.filter(usuario=usuario, mes=mes, anio=anio)
        .select_related("categoria", "rubro")
        .order_by("-fecha")
    )

    gastos_por_categoria = (
        gastos.values("categoria__nombre", "categoria__color")
        .annotate(total=Sum("monto"))
        .order_by("-total")
    )

    ingresos_nomina = RegistroNomina.objects.filter(usuario=usuario, mes=mes, anio=anio)
    otros_ingresos = OtroIngreso.objects.filter(usuario=usuario, mes=mes, anio=anio)

    creditos_activos = Credito.objects.filter(usuario=usuario, activo=True)
    tarjetas_activas = TarjetaCredito.objects.filter(usuario=usuario, activa=True)

    provisiones_activas = Provision.objects.filter(
        usuario=usuario, activa=True
    ).order_by("-fecha_pago")

    ctx = {
        "usuario": usuario,
        "mes": mes,
        "anio": anio,
        "mes_nombre": MESES_NOMBRE[mes],
        "indicadores": indicadores,
        "gastos": gastos,
        "gastos_por_categoria": gastos_por_categoria,
        "ingresos_nomina": ingresos_nomina,
        "otros_ingresos": otros_ingresos,
        "creditos_activos": creditos_activos,
        "tarjetas_activas": tarjetas_activas,
        "provisiones_activas": provisiones_activas,
        "generado_en": _marca_temporal(),
        "total_ingresos": indicadores["ingreso_neto"],
        "total_gastos": indicadores["gastos_totales"],
        "ahorro_neto": indicadores["ahorro_neto"],
        "tasa_ahorro": indicadores["tasa_ahorro"],
    }

    html_string = render_to_string("reportes/reporte_mensual.html", ctx)
    pdf_file = BytesIO()
    HTML(string=html_string).write_pdf(pdf_file)
    pdf_file.seek(0)

    filename = _nombre_archivo("reporte_mensual", "pdf", usuario.username, mes, anio)
    response = HttpResponse(pdf_file.read(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def exportar_csv(usuario, mes: int, anio: int) -> HttpResponse:
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = (
        f'attachment; filename="{_nombre_archivo("exportacion", "csv", usuario.username, mes, anio)}"'
    )

    writer = csv.writer(response)
    writer.writerow(["Finanzas Hogar — Exportación", f"{MESES_NOMBRE[mes]} {anio}"])
    writer.writerow([f"Generado: {_marca_temporal()}"])
    writer.writerow([])

    writer.writerow(["INGRESOS"])
    writer.writerow(["Tipo", "Concepto", "Valor COP"])
    for n in RegistroNomina.objects.filter(usuario=usuario, mes=mes, anio=anio):
        writer.writerow(["Nómina", f"{n.mes}/{n.anio}", str(n.salario_neto)])
    for o in OtroIngreso.objects.filter(usuario=usuario, mes=mes, anio=anio):
        writer.writerow(["Otro ingreso", o.tipo, str(o.monto)])

    writer.writerow([])
    writer.writerow(["GASTOS"])
    writer.writerow(["Categoría", "Rubro", "Descripción", "Fecha", "Valor COP"])
    for g in Gasto.objects.filter(usuario=usuario, mes=mes, anio=anio).select_related(
        "categoria", "rubro"
    ):
        writer.writerow(
            [g.categoria.nombre, g.rubro.nombre, g.descripcion, g.fecha, str(g.monto)]
        )

    writer.writerow([])
    writer.writerow(["DEUDAS — CRÉDITOS"])
    writer.writerow(["Crédito", "Cuota mensual", "Saldo restante"])
    for c in Credito.objects.filter(usuario=usuario, activo=True):
        cuotas_pagadas = CuotaCredito.objects.filter(credito=c, pagada=True).count()
        total_cuotas = CuotaCredito.objects.filter(credito=c).count()
        saldo_restante = CuotaCredito.objects.filter(credito=c, pagada=False).aggregate(
            s=Sum("saldo_capital")
        )["s"] or Decimal("0")
        writer.writerow([c.nombre, str(c.cuota_mensual or 0), str(saldo_restante)])

    writer.writerow([])
    writer.writerow(["DEUDAS — TARJETAS DE CRÉDITO"])
    writer.writerow(["Tarjeta", "Saldo actual", "Límite", "Disponible"])
    for t in TarjetaCredito.objects.filter(usuario=usuario, activa=True):
        writer.writerow(
            [t.nombre, str(t.saldo_actual), str(t.limite), str(t.disponible)]
        )

    writer.writerow([])
    writer.writerow(["PROVISIONES ACTIVAS"])
    writer.writerow(["Concepto", "Monto total", "Ahorro acumulado", "Fecha pago"])
    for p in Provision.objects.filter(usuario=usuario, activa=True):
        writer.writerow(
            [p.concepto, str(p.monto_total), str(p.ahorro_acumulado), p.fecha_pago]
        )

    return response


def exportar_excel(usuario, mes: int, anio: int) -> HttpResponse:
    wb = openpyxl.Workbook()

    # --- Hoja: Resumen ---
    ws_resumen = wb.active
    ws_resumen.title = "Resumen"
    ws_resumen.merge_cells("A1:D1")
    ws_resumen["A1"] = f"Finanzas Hogar — {MESES_NOMBRE[mes]} {anio}"
    ws_resumen["A1"].font = Font(bold=True, size=14)
    ws_resumen["A2"] = f"Generado: {_marca_temporal()}"
    ws_resumen["A2"].font = Font(italic=True, size=10, color="666666")

    indicadores = calcular_indicadores_mes(usuario, mes, anio)
    headers_resumen = ["Indicador", "Valor", "Semáforo", "Meta"]
    _estilo_cabecera(ws_resumen, headers_resumen, row=4)
    datos_resumen = [
        ("Ingreso neto", f'${indicadores["ingreso_neto"]:,.0f}', "", ""),
        ("Gastos totales", f'${indicadores["gastos_totales"]:,.0f}', "", ""),
        ("Ahorro neto", f'${indicadores["ahorro_neto"]:,.0f}', "", ""),
        (
            "Tasa de ahorro",
            f'{indicadores["tasa_ahorro"]:.1f}%',
            indicadores.get("semaforo_ahorro", ""),
            f"{usuario.meta_tasa_ahorro}%",
        ),
        (
            "Ratio endeudamiento",
            f'{indicadores["ratio_endeudamiento"]:.1f}%',
            indicadores.get("semaforo_endeudamiento", ""),
            "< 30%",
        ),
        (
            "Cobertura emergencia",
            f'{indicadores["cobertura_emergencia"]:.1f} meses',
            indicadores.get("semaforo_emergencia", ""),
            "≥ 3 meses",
        ),
        (
            "Presión gastos fijos",
            f'{indicadores["presion_gastos_fijos"]:.1f}%',
            "",
            "< 50%",
        ),
    ]
    for i, (k, v, s, m) in enumerate(datos_resumen, 5):
        ws_resumen.cell(row=i, column=1, value=k).border = BORDER_THIN
        ws_resumen.cell(row=i, column=2, value=v).border = BORDER_THIN
        ws_resumen.cell(row=i, column=3, value=s).border = BORDER_THIN
        ws_resumen.cell(row=i, column=4, value=m).border = BORDER_THIN

    # --- Hoja: Ingresos ---
    ws_ingresos = wb.create_sheet("Ingresos")
    headers_ing = ["Tipo", "Concepto", "Valor COP"]
    _estilo_cabecera(ws_ingresos, headers_ing)
    row = 2
    for n in RegistroNomina.objects.filter(usuario=usuario, mes=mes, anio=anio):
        ws_ingresos.cell(row=row, column=1, value="Nómina").border = BORDER_THIN
        ws_ingresos.cell(row=row, column=2, value=f"{n.mes}/{n.anio}").border = (
            BORDER_THIN
        )
        ws_ingresos.cell(row=row, column=3, value=float(n.salario_neto)).border = (
            BORDER_THIN
        )
        ws_ingresos.cell(row=row, column=3).number_format = "#,##0"
        row += 1
    for o in OtroIngreso.objects.filter(usuario=usuario, mes=mes, anio=anio):
        ws_ingresos.cell(row=row, column=1, value="Otro ingreso").border = BORDER_THIN
        ws_ingresos.cell(row=row, column=2, value=o.tipo).border = BORDER_THIN
        ws_ingresos.cell(row=row, column=3, value=float(o.monto)).border = BORDER_THIN
        ws_ingresos.cell(row=row, column=3).number_format = "#,##0"
        row += 1
    _autoajuste_columnas(ws_ingresos)

    # --- Hoja: Gastos ---
    ws_gastos = wb.create_sheet("Gastos")
    headers_gas = ["Categoría", "Rubro", "Descripción", "Fecha", "Valor COP"]
    _estilo_cabecera(ws_gastos, headers_gas)
    row = 2
    for g in Gasto.objects.filter(usuario=usuario, mes=mes, anio=anio).select_related(
        "categoria", "rubro"
    ):
        ws_gastos.cell(row=row, column=1, value=g.categoria.nombre).border = BORDER_THIN
        ws_gastos.cell(row=row, column=2, value=g.rubro.nombre).border = BORDER_THIN
        ws_gastos.cell(row=row, column=3, value=g.descripcion).border = BORDER_THIN
        ws_gastos.cell(row=row, column=4, value=g.fecha.isoformat()).border = (
            BORDER_THIN
        )
        ws_gastos.cell(row=row, column=5, value=float(g.monto)).border = BORDER_THIN
        ws_gastos.cell(row=row, column=5).number_format = "#,##0"
        row += 1
    _autoajuste_columnas(ws_gastos)

    # --- Hoja: Deudas ---
    ws_deudas = wb.create_sheet("Deudas")
    headers_deu = ["Tipo", "Nombre", "Cuota mensual", "Saldo actual", "Límite/Plazo"]
    _estilo_cabecera(ws_deudas, headers_deu)
    row = 2
    for c in Credito.objects.filter(usuario=usuario, activo=True):
        saldo = CuotaCredito.objects.filter(credito=c, pagada=False).aggregate(
            s=Sum("saldo_capital")
        )["s"] or Decimal("0")
        ws_deudas.cell(row=row, column=1, value="Crédito").border = BORDER_THIN
        ws_deudas.cell(row=row, column=2, value=c.nombre).border = BORDER_THIN
        ws_deudas.cell(row=row, column=3, value=float(c.cuota_mensual or 0)).border = (
            BORDER_THIN
        )
        ws_deudas.cell(row=row, column=3).number_format = "#,##0"
        ws_deudas.cell(row=row, column=4, value=float(saldo)).border = BORDER_THIN
        ws_deudas.cell(row=row, column=4).number_format = "#,##0"
        ws_deudas.cell(row=row, column=5, value=f"{c.plazo_meses} meses").border = (
            BORDER_THIN
        )
        row += 1
    for t in TarjetaCredito.objects.filter(usuario=usuario, activa=True):
        ws_deudas.cell(row=row, column=1, value="Tarjeta").border = BORDER_THIN
        ws_deudas.cell(row=row, column=2, value=t.nombre).border = BORDER_THIN
        ws_deudas.cell(
            row=row, column=3, value=float(t.cuota_minima_mes or 0)
        ).border = BORDER_THIN
        ws_deudas.cell(row=row, column=3).number_format = "#,##0"
        ws_deudas.cell(row=row, column=4, value=float(t.saldo_actual)).border = (
            BORDER_THIN
        )
        ws_deudas.cell(row=row, column=4).number_format = "#,##0"
        ws_deudas.cell(
            row=row, column=5, value=f"Límite: ${float(t.limite):,.0f}"
        ).border = BORDER_THIN
        row += 1
    _autoajuste_columnas(ws_deudas)

    # --- Hoja: Provisiones ---
    ws_prov = wb.create_sheet("Provisiones")
    headers_prov = ["Concepto", "Monto total", "Ahorro acumulado", "Fecha pago"]
    _estilo_cabecera(ws_prov, headers_prov)
    row = 2
    for p in Provision.objects.filter(usuario=usuario, activa=True):
        ws_prov.cell(row=row, column=1, value=p.concepto).border = BORDER_THIN
        ws_prov.cell(row=row, column=2, value=float(p.monto_total)).border = BORDER_THIN
        ws_prov.cell(row=row, column=2).number_format = "#,##0"
        ws_prov.cell(row=row, column=3, value=float(p.ahorro_acumulado)).border = (
            BORDER_THIN
        )
        ws_prov.cell(row=row, column=3).number_format = "#,##0"
        ws_prov.cell(row=row, column=4, value=p.fecha_pago.isoformat()).border = (
            BORDER_THIN
        )
        row += 1
    _autoajuste_columnas(ws_prov)

    # --- Hoja: Fondo de emergencia ---
    ws_fondo = wb.create_sheet("Fondo Emergencia")
    fondo = FondoEmergencia.objects.filter(usuario=usuario).first()
    ws_fondo["A1"] = "Saldo actual"
    ws_fondo["B1"] = float(fondo.saldo_actual) if fondo else 0
    ws_fondo["B1"].number_format = "#,##0"
    ws_fondo["A2"] = "Cobertura"
    ws_fondo["B2"] = f'{indicadores["cobertura_emergencia"]:.1f} meses'
    ws_fondo["A3"] = "Gasto esencial mensual"
    ws_fondo["B3"] = float(indicadores.get("gasto_esencial", 0))
    ws_fondo["B3"].number_format = "#,##0"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = (
        f'attachment; filename="{_nombre_archivo("exportacion", "xlsx", usuario.username, mes, anio)}"'
    )
    wb.save(response)
    return response
