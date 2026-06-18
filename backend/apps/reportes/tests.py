import pytest
from io import BytesIO
from decimal import Decimal
from datetime import date
from django.urls import reverse
from django.test import Client
from apps.accounts.models import UserProfile
from apps.gastos.models import Gasto, Categoria, Rubro
from apps.ingresos.models import RegistroNomina
from apps.deudas.models import Credito, TarjetaCredito
from apps.provisiones.models import Provision
from .services import _marca_temporal, _nombre_archivo


@pytest.mark.django_db
class TestServicesReportes:

    def _crear_usuario(self):
        return UserProfile.objects.create_user(
            username="testuser",
            email="test@example.co",
            password="pass123",
            nombre_completo="Test Usuario",
            meta_tasa_ahorro=Decimal("20"),
        )

    def test_marca_temporal_retorna_string(self):
        ts = _marca_temporal()
        assert isinstance(ts, str)
        assert "/" in ts

    def test_nombre_archivo_formato(self):
        name = _nombre_archivo("reporte_mensual", "pdf", "testuser", 6, 2026)
        assert name.startswith("reporte_mensual_testuser_202606_")
        assert name.endswith(".pdf")

    def test_nombre_archivo_excel(self):
        name = _nombre_archivo("exportacion", "xlsx", "testuser", 12, 2026)
        assert name.endswith(".xlsx")

    def test_generar_pdf_mes_retorna_pdf(self, db):
        from .services import generar_pdf_mes

        user = self._crear_usuario()
        hoy = date.today()
        response = generar_pdf_mes(user, hoy.month, hoy.year)
        assert response.status_code == 200
        assert response["Content-Type"] == "application/pdf"
        assert "Content-Disposition" in response
        assert response["Content-Disposition"].startswith(
            'attachment; filename="reporte_mensual_'
        )

    def test_generar_pdf_con_datos(self, db):
        from .services import generar_pdf_mes

        user = self._crear_usuario()
        hoy = date.today()

        cat = Categoria.objects.create(
            nombre="TestCat", color="#FF0000", es_esencial=False, visible=True
        )
        rubro = Rubro.objects.create(categoria=cat, nombre="TestRubro", visible=True)

        Gasto.objects.create(
            usuario=user,
            categoria=cat,
            rubro=rubro,
            monto=Decimal("150000"),
            fecha=hoy,
            metodo_pago="efectivo",
            tipo="variable",
            mes=hoy.month,
            anio=hoy.year,
        )
        RegistroNomina.objects.create(
            usuario=user,
            salario_bruto=Decimal("2000000"),
            deduccion_salud=Decimal("80000"),
            deduccion_pension=Decimal("80000"),
            salario_neto=Decimal("1840000"),
            neto_con_auxilio=Decimal("1840000"),
            mes=hoy.month,
            anio=hoy.year,
        )

        response = generar_pdf_mes(user, hoy.month, hoy.year)
        assert response.status_code == 200
        assert response["Content-Type"] == "application/pdf"

    def test_exportar_csv_sin_datos(self, db):
        from .services import exportar_csv

        user = self._crear_usuario()
        hoy = date.today()
        response = exportar_csv(user, hoy.month, hoy.year)
        assert response.status_code == 200
        assert response["Content-Type"] == "text/csv; charset=utf-8"
        content = response.content.decode("utf-8")
        assert "Finanzas Hogar" in content

    def test_exportar_csv_con_datos(self, db):
        from .services import exportar_csv

        user = self._crear_usuario()
        hoy = date.today()

        cat = Categoria.objects.create(
            nombre="Comida", color="#FF0000", es_esencial=True, visible=True
        )
        rubro = Rubro.objects.create(categoria=cat, nombre="Mercado", visible=True)
        Gasto.objects.create(
            usuario=user,
            categoria=cat,
            rubro=rubro,
            monto=Decimal("250000"),
            fecha=hoy,
            metodo_pago="efectivo",
            tipo="variable",
            mes=hoy.month,
            anio=hoy.year,
        )

        response = exportar_csv(user, hoy.month, hoy.year)
        assert response.status_code == 200
        content = response.content.decode("utf-8")
        assert "GASTOS" in content
        assert "Comida" in content
        assert "Mercado" in content

    def test_exportar_excel_retorna_xlsx(self, db):
        from .services import exportar_excel

        user = self._crear_usuario()
        hoy = date.today()
        response = exportar_excel(user, hoy.month, hoy.year)
        assert response.status_code == 200
        assert (
            response["Content-Type"]
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        assert response["Content-Disposition"].startswith(
            'attachment; filename="exportacion_'
        )

    def test_exportar_excel_con_datos(self, db):
        from .services import exportar_excel

        user = self._crear_usuario()
        hoy = date.today()

        cat = Categoria.objects.create(
            nombre="Transporte", color="#00FF00", es_esencial=False, visible=True
        )
        rubro = Rubro.objects.create(categoria=cat, nombre="Gasolina", visible=True)
        Gasto.objects.create(
            usuario=user,
            categoria=cat,
            rubro=rubro,
            monto=Decimal("120000"),
            fecha=hoy,
            metodo_pago="efectivo",
            tipo="variable",
            mes=hoy.month,
            anio=hoy.year,
        )
        Provision.objects.create(
            usuario=user,
            concepto="SOAT",
            monto_total=Decimal("500000"),
            ahorro_acumulado=Decimal("100000"),
            fecha_pago=hoy,
            activa=True,
        )

        response = exportar_excel(user, hoy.month, hoy.year)
        assert response.status_code == 200

        wb = openpyxl_load(BytesIO(response.content))
        assert "Resumen" in wb.sheetnames
        assert "Gastos" in wb.sheetnames
        assert "Provisiones" in wb.sheetnames

        ws_gastos = wb["Gastos"]
        valores = []
        for row in ws_gastos.iter_rows(min_row=2, values_only=True):
            if row[0]:
                valores.append(row[0])
        assert "Transporte" in valores


def openpyxl_load(io_bytes):
    import openpyxl

    return openpyxl.load_workbook(io_bytes)


@pytest.mark.django_db
class TestViewsReportes:

    def _crear_usuario(self):
        return UserProfile.objects.create_user(
            username="testuser",
            email="test@example.co",
            password="pass123",
            nombre_completo="Test",
        )

    def test_reporte_pdf_requiere_auth(self):
        response = Client().get(reverse("reportes:reporte_pdf"))
        assert response.status_code == 302

    def test_exportar_csv_requiere_auth(self):
        response = Client().get(reverse("reportes:exportar_csv"))
        assert response.status_code == 302

    def test_exportar_excel_requiere_auth(self):
        response = Client().get(reverse("reportes:exportar_excel"))
        assert response.status_code == 302

    def test_reporte_pdf_autenticado(self, db):
        user = self._crear_usuario()
        client = Client()
        client.force_login(user)
        response = client.get(reverse("reportes:reporte_pdf"))
        assert response.status_code == 200
        assert response["Content-Type"] == "application/pdf"

    def test_exportar_csv_autenticado(self, db):
        user = self._crear_usuario()
        client = Client()
        client.force_login(user)
        response = client.get(reverse("reportes:exportar_csv"))
        assert response.status_code == 200
        assert "text/csv" in response["Content-Type"]

    def test_exportar_excel_autenticado(self, db):
        user = self._crear_usuario()
        client = Client()
        client.force_login(user)
        response = client.get(reverse("reportes:exportar_excel"))
        assert response.status_code == 200
        assert "spreadsheetml" in response["Content-Type"]

    def test_reporte_pdf_con_mes_anio(self, db):
        user = self._crear_usuario()
        client = Client()
        client.force_login(user)
        response = client.get(reverse("reportes:reporte_pdf") + "?mes=6&anio=2026")
        assert response.status_code == 200

    def test_exportar_csv_con_mes_anio(self, db):
        user = self._crear_usuario()
        client = Client()
        client.force_login(user)
        response = client.get(reverse("reportes:exportar_csv") + "?mes=12&anio=2026")
        assert response.status_code == 200
